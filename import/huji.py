import openpyxl
import time
import os
from mariadb.mariadb_manager import MariadbManager
import datetime

def parse_huji_excel(excel_full_path, mariadb_conn, insert_sql_stmt):

    excel_book = openpyxl.load_workbook(excel_full_path, read_only=True)
    excel_sheet = excel_book.worksheets[0]

    print("总行数：" + str(excel_sheet.max_row))
    print("总列数：" + str(excel_sheet.max_column))

    phone_is_null_num = 0

    logfile_path = "D:\log.txt"
    inserted_num = 0
    datarow_list = []
    row_num = 0
    for cur_row in excel_sheet.rows:
        row_num = row_num + 1
        # print(row_num)
        if row_num == 1:
            continue

        """"
        temp_stat_time = cur_row[1].value.strip()   # cur_row[1]  统计时间
        temp_stat_no = cur_row[3].value.strip()     # cur_row[3]  编码编号
        temp_birth_date = cur_row[4].value.strip()  # cur_row[4]  出生日期
        temp_sex = cur_row[5].value.strip()         # cur_row[5]  性别
        temp_id_no = cur_row[6].value.strip()       # cur_row[6]  身份证号
        temp_name = cur_row[7].value.strip()        # cur_row[7]  姓名
        temp_phone = str(cur_row[8].value).strip()       # cur_row[8]  电话号码
        temp_address = cur_row[9].value.strip()     # cur_row[9]  地址
        """
        temp_stat_time = cell_value_to_str(cur_row, cur_row[1].value)   # cur_row[1]  统计时间
        temp_stat_no = cell_value_to_str(cur_row, cur_row[3].value)     # cur_row[3]  编码编号
        temp_birth_date = cell_value_to_str(cur_row, cur_row[4].value)  # cur_row[4]  出生日期
        temp_sex = cell_value_to_str(cur_row, cur_row[5].value)         # cur_row[5]  性别
        temp_id_no = cell_value_to_str(cur_row, cur_row[6].value)       # cur_row[6]  身份证号
        temp_name = cell_value_to_str(cur_row, cur_row[7].value)        # cur_row[7]  姓名
        temp_phone = cell_value_to_str(cur_row, cur_row[8].value)       # cur_row[8]  电话号码
        temp_address = cell_value_to_str(cur_row, cur_row[9].value)     # cur_row[9]  地址

        if temp_id_no.find("'") >= 0:
            temp_id_no = temp_id_no.replace("'", "")

        if temp_phone.find("'") >= 0:
            temp_phone = temp_phone.replace("'", "")

        if len(temp_id_no) > 30:
            print(str(row_num) + "身份证号码长度大于30:%s" % temp_id_no)
        if len(temp_phone) > 30:
            print(str(row_num) + " " + temp_id_no + "手机号码长度大于30:%s" % temp_phone)
        if len(temp_address) > 300:
            print(str(row_num) + " " + temp_id_no + "地址长度大于30:%s" % temp_address)
        if len(temp_phone) == 0:
            phone_is_null_num = phone_is_null_num +1
            # print(temp_stat_time + "用户:%s 的电话号码为空，本条数据跳过" % temp_name)
            continue

        mysql_values_tuple = (temp_stat_time, temp_stat_no, temp_birth_date,
                              temp_sex, temp_id_no, temp_name, temp_phone, temp_address)
        datarow_list.append(mysql_values_tuple)
        if len(datarow_list) == 10000:
            temp_num = batch_insert_data(mariadb_conn, insert_sql_stmt, datarow_list)
            inserted_num = inserted_num + temp_num
            datarow_list.clear()
            print("共计插入数据%d行" % inserted_num)

    if len(datarow_list) > 0:
        temp_count = batch_insert_data(mariadb_conn, insert_sql_stmt, datarow_list)
        inserted_num = inserted_num + temp_count
        print("共计插入数据%d行" % inserted_num)
    print("文件入库完成，共入库数据"+str(inserted_num)+"行")

    with open(logfile_path, "a") as file:
        file.write(excel_full_path + "  文件入库完成，共入库数据"+str(inserted_num)+"行,因电话号码空没有插入的数据条数为：%d" % phone_is_null_num + "\n")


def cell_value_to_str(cur_row, cell_value):
    return_value = ""
    try:
        if cell_value is None:
            return_value = ""
        elif type(cell_value) == int:
            return_value = str(cell_value).strip()
        elif type(cell_value) == str:
            return_value = cell_value.strip()
        elif type(cell_value) == float:
            return_value = str(cell_value).strip()
        elif type(cell_value) == datetime.datetime:
            return_value = cell_value.strftime("%Y/%m/%d")
        else:
            print("Cell:"+cur_row[6].value+",未知的cell数据类型:" + type(cell_value))

    except Exception as ex:
        print("--------"+ex)

    return return_value


def batch_insert_data(mariadb_conn, sql_stmt, values_list):

    begin_time = time.time()
    table_cursor = mariadb_conn.cursor()
    total_count = 0
    try:
        total_count = table_cursor.executemany(sql_stmt, values_list)
        mariadb_conn.commit()
    except Exception as ex:
        print(ex)
        mariadb_conn.rollback()
    table_cursor.close()
    end_time = time.time()
    spend_time = end_time-begin_time
    print("本次插入数据%d条" % total_count+"  共计花费时间:%s秒" % format(spend_time, '0.2f'))
    return total_count


def check_huji_file_format(full_path, column_header_list):

    excel_book = openpyxl.load_workbook(full_path, read_only=True)
    excel_sheet = excel_book.worksheets[0]
    if excel_sheet.max_row > 1:
        is_same = True
        same_num = 0
        diff_column_name = ""
        temp = ""
        for i in range(len(column_header_list)):
            if excel_sheet.cell(1, i+1).value == column_header_list[i]:
                temp = temp + excel_sheet.cell(1, i + 1).value + ","
                same_num = same_num + 1
            else:
                is_same = False
                diff_column_name = diff_column_name
                break

        if is_same is False:
            print("文件:"+full_path+" 的列头与标准列头不一致，请检查")
            print(diff_column_name)
        else:
            print("文件:" + full_path + " 的格式无误")

        temp_id_no = excel_sheet.cell(2, 7).value.strip()  # cur_row[7]  身份证号
        temp_name = excel_sheet.cell(2, 8).value.strip()  # cur_row[8]  姓名
        temp_phone = excel_sheet.cell(2, 9).value.strip()  # cur_row[9]  电话号码
        if len(temp_id_no) != 15:
            if len(temp_id_no) != 18:
                print("身份证格式有误:%s" % temp_id_no)

    else:
        print("文件:" + full_path + " 的行数为空，请检查！")


def parse_all_huji_data():

    logfile_path = "D:\log.txt"
    with open(logfile_path, "w", encoding="utf-8") as file:
        print("初始化日志文件")

    mariadb_manager = MariadbManager("127.0.0.1", 3306, "privacydata", "root", "pmo@2016",  charset="utf8mb4")
    mariadb_manager.open_connect()
    insert_sql_stmt = "INSERT INTO HUJI (STAT_TIME, STAT_NO, BIRTH_DATE, SEX, ID_NO, NAME, PHONE, ADDRESS)"
    insert_sql_stmt = insert_sql_stmt + " VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
    column_header = ["所属户籍站", "统计时间", "居住地址", "编码编号", "出生日期", "性别", "身份证", "名字", "手机号", "地址", "编码", "编码"]
    rootdir = r"D:\downloads\privacydata\huji"
    list = os.listdir(rootdir)  # 列出文件夹下所有的目录与文件
    file_num = 0
    for i in range(0, len(list)):
        file_path = os.path.join(rootdir, list[i])
        if os.path.isfile(file_path):
            file_num = file_num + 1
            print("开始第" + str(file_num) + "文件:" + file_path + "！")
            # if file_path == r"D:\downloads\privacydata\huji\1 - 副本 (18).xlsx":

            # check_huji_file_format(file_path, column_header)
            parse_huji_excel(file_path, mariadb_manager.connect, insert_sql_stmt)



def main():

    try:
        parse_all_huji_data()
    except Exception as ex:
        print(ex)


if __name__ == '__main__':

    start_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    print('Beginning ! Begin:' + start_time)
    main()
    end_time = time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())
    print('\nEnd:' + end_time)
